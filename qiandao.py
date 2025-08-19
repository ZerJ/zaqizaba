# -*- coding: utf-8 -*-
# qiandao_incremental_excel_v3.py

import requests
import time
import hmac
import hashlib
import base64
from urllib.parse import urlparse, parse_qsl, quote
import logging
from openpyxl import Workbook, load_workbook
from pathlib import Path

# === 日志配置 ===
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("qiandao_scrape.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)

# === SKEY 配置 ===
SKEYS = {
    "production": "8eVqsqhBdEKxeULQ1M72yrZB5HET9fx0",
    "development": "hX3fT6ZaCyxtLiogRIO1NUKADxBUNHQL",
}

# === Excel 文件 ===
ENTRY_ID_FILE = "entry_ids.xlsx"
DETAIL_FILE = "千岛.xlsx"
FAILED_FILE = "failed.xlsx"


# ---------------------------
# HMAC 签名和请求
# ---------------------------
def get_env(url: str) -> str:
    host = urlparse(url).netloc.lower()
    if host.startswith(("dev-", "local")) or host.startswith("test-"):
        return "development"
    return "production"


def canonical_query(qs: str, extra: dict | None = None) -> str:
    pairs = []
    if qs:
        pairs.extend(parse_qsl(qs, keep_blank_values=True, strict_parsing=False))
    if extra:
        for k, v in extra.items():
            if isinstance(v, (list, tuple)):
                for item in v:
                    pairs.append((str(k), "" if item is None else str(item)))
            else:
                pairs.append((str(k), "" if v is None else str(v)))
    pairs.sort(key=lambda kv: (kv[0], kv[1]))
    return "&".join(f"{quote(k, safe='')}={v}" for k, v in pairs)


def make_sign(url: str, timestamp_ms: int, request_id: str = "", extra_params: dict | None = None) -> str:
    parsed = urlparse(url)
    path = parsed.path
    cq = canonical_query(parsed.query, extra_params)
    base_string = f"{path}{cq}{request_id}{timestamp_ms}"
    env = get_env(url)
    skey = SKEYS[env]
    digest = hmac.new(skey.encode("utf-8"),
                      base_string.encode("utf-8"),
                      hashlib.sha256).digest()
    return base64.b64encode(digest.hex().encode("ascii")).decode("ascii")


def build_headers(url: str, request_id: str = "", timestamp_ms: int | None = None,
                  extra_params: dict | None = None) -> dict:
    if timestamp_ms is None:
        timestamp_ms = int(time.time() * 1000)
    sign = make_sign(url, timestamp_ms, request_id=request_id, extra_params=extra_params)
    headers = {
        "x-request-timestamp": str(timestamp_ms),
        "x-request-sign": sign,
        "x-request-sign-type": "HMAC_SHA256",
        "x-request-sign-version": "v1",
        "x-request-package-sign-version": "0.0.2",
        "Host": urlparse(url).netloc,
        "x-echoing-env": "test-z",
        "x-request-package-id": "1000",
        "x-request-circle": "chaowanzu",
        "x-request-version": "2.123.2",
        "xweb_xhr": "1",
        "x-package-id": "1000",
        "x-device-id": "f3bf1984-dac2-47e4-b7fd-31658d3ba637",
        "content-type": "application/json",
        "accept": "*/*",
    }
    return headers


def request_with_retry(method, url, headers=None, json_data=None, retries=3, delay=2.0):
    for attempt in range(1, retries + 1):
        try:
            if method.lower() == "post":
                response = requests.post(url, headers=headers, json=json_data, timeout=10)
            else:
                response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            return response
        except Exception as e:
            logging.warning(f"请求 {url} 第 {attempt} 次失败: {e}")
            if attempt < retries:
                logging.info(f"{delay} 秒后重试...")
                time.sleep(delay)
            else:
                logging.error(f"请求 {url} 达到最大重试次数，跳过。")
                return None


# ---------------------------
# 解析 simpleInfo 数据
# ---------------------------
def parse_simple_info(data: dict, entry_id: str):
    header = data.get("data").get("header", {})
    content = data.get("data").get("content", {})

    card_name = header.get("name", "")
    description = header.get("description", "")
    props = {p["propertyName"]: p["dataValue"] for p in header.get("baseKeyPropertyInfos", [])}
    print(props)
    print(header)
    alias= header.get("description", "")

    artist = ""
    group = ""
    for ch in content.get("characters", []):
        if ch.get("type") == "tag":
            if "islandId" in ch:
                group = ch.get("name", "")
            else:
                artist = ch.get("name", "")

    return {
        "entryId": entry_id,
        "小卡名称": card_name,
        "别名": alias,
        "艺人": artist,
        "团体": group,
        "发售平台": props.get("发售平台", ""),
        "分类": props.get("分类", ""),
        "发售日期": props.get("发售日期", ""),
        "发售地区": props.get("发售地区", ""),
        "艺人公司": props.get("艺人公司", "")
    }


# ---------------------------
# Feed 抓取
# ---------------------------
def fetch_all_entry_ids(limit=20, delay=1.0):
    offset = 0
    all_entry_ids = []
    while True:
        url = "https://api.qiandao.cn/treasure/spus/feed"
        payload = {"limit": limit, "offset": offset, "orderBy": "waterfallScoreDesc", "andTagIds": ["1512146"]}
        headers = build_headers(url)
        response = request_with_retry("post", url, headers=headers, json_data=payload)
        if response is None:
            time.sleep(delay)
            continue
        data = response.json()
        items = data.get("data", {}).get("list", [])
        if not items:
            break
        for item in items:
            entry_id = item.get("id")
            if entry_id and entry_id not in all_entry_ids:
                all_entry_ids.append(entry_id)
        offset += limit
        logging.info(f"已抓取 feed {offset} 条, 当前总 entryId 数: {len(all_entry_ids)}")
        time.sleep(delay)
    return all_entry_ids


def fetch_entry_detail(entry_id: str, delay: float = 1.0):
    url = f"https://api.qiandao.com/treasure/flora/v2/spu/simpleInfo?entryId={entry_id}"
    headers = build_headers(url)
    fail_count = 0
    while True:
        response = request_with_retry("get", url, headers=headers)
        if response is None:
            fail_count += 1
            time.sleep(delay * (1.5 ** fail_count))
            if fail_count > 3:
                save_failed_entry_id_to_excel(entry_id)
                return None
            continue
        time.sleep(delay)

        return response.json()


# ---------------------------
# Excel 工具
# ---------------------------
class ExcelWriter:
    def __init__(self, file_path, headers=None, sheet_name=None):
        self.file_path = file_path
        self.headers = headers
        self.sheet_name = sheet_name
        self.existing_ids = set()
        if Path(file_path).exists():
            self.wb = load_workbook(file_path)
            self.ws = self.wb.active
            self.existing_ids = {row[0].value for row in self.ws.iter_rows(min_row=2, values_only=True) if row[0].value}
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            if sheet_name:
                self.ws.title = sheet_name
            if headers:
                self.ws.append(headers)

    def append_row(self, row, unique_id=None):
        if unique_id and unique_id in self.existing_ids:
            return False
        self.ws.append(row)
        if unique_id:
            self.existing_ids.add(unique_id)
        return True

    def save(self):
        self.wb.save(self.file_path)
        logging.info(f"已保存 {self.file_path}")


# ---------------------------
# 增量保存 entryId
# ---------------------------
def load_existing_entry_ids_from_file(file_path=ENTRY_ID_FILE):
    if Path(file_path).exists():
        wb = load_workbook(file_path)
        ws = wb.active
        return {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    return set()


def save_entry_ids_to_excel_incremental(entry_ids, file_path=ENTRY_ID_FILE):
    existing_ids = load_existing_entry_ids_from_file(file_path)
    new_ids = [eid for eid in entry_ids if eid not in existing_ids]
    if not new_ids:
        logging.info("没有新的 entryId 需要保存")
        return
    if Path(file_path).exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "EntryIDs"
        ws.append(["entryId"])
    for eid in new_ids:
        ws.append([eid])
    wb.save(file_path)
    logging.info(f"已追加 {len(new_ids)} 个新 entryId 到 {file_path}")


# ---------------------------
# 失败记录
# ---------------------------
def save_failed_entry_id_to_excel(entry_id, file_path=FAILED_FILE):
    if Path(file_path).exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "FailedEntryIDs"
        ws.append(["entryId"])
    ws.append([entry_id])
    wb.save(file_path)
    logging.warning(f"已保存失败 entryId={entry_id} 到 {file_path}")


# ---------------------------
# 主流程
# ---------------------------
def incremental_fetch(entry_delay=1.0, feed_delay=1.5):
    logging.info("=== 获取 feed 中的所有 entryId ===")
    # entry_ids = fetch_all_entry_ids(delay=feed_delay)
    # logging.info(f"共获取 {len(entry_ids)} 个 entryId")
    entry_ids=load_existing_entry_ids_from_file()
    # 增量保存 entryId
    # save_entry_ids_to_excel_incremental(entry_ids)
    print(entry_ids)
    # 加载详情已有记录
    detail_headers = ["entryId", "小卡名称", "别名", "艺人", "团体", "发售平台", "分类", "发售日期",
                      "发售地区", "艺人公司", "描述"]
    detail_writer = ExcelWriter(DETAIL_FILE, headers=detail_headers, sheet_name="Details")

    for entry_id in entry_ids:
        if entry_id in detail_writer.existing_ids:
            continue
        data = fetch_entry_detail(entry_id, delay=entry_delay)
        if data:
            record = parse_simple_info(data, entry_id)

            detail_writer.append_row([record.get(h, "") for h in detail_headers], unique_id=entry_id)
            logging.info(f"抓取 entryId={entry_id} 成功: {record['小卡名称']}")
        else:
            logging.error(f"抓取 entryId={entry_id} 失败，已跳过")

    detail_writer.save()


if __name__ == "__main__":
    incremental_fetch(entry_delay=5.0, feed_delay=6.0)
